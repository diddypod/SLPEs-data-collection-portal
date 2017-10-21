import os
from openpyxl import load_workbook
import pymysql.cursors
import tkinter as tk
from tkinter import *

typs = [['engg','Engineering'], ['pnp','Printing & Publication'], ['enr','Energy'],
       ['phrm','Pharmaceuticals'], ['txt','Textile'], ['trad','Trading'], ['manu','Manufacturing'],
       ['chem','Chemicals'], ['agro','Agro-Industries'], ['tran','Transport'],
       ['misc','Development of Service & Misc. Sectors']]
year=''

def getList():
    global year
    yr = enty.get()
    if yr != '':
        year=yr
    else:
        tk.messagebox.showerror(title="Error",message="No year entered.\nPlease enter years.")
    root.destroy()
root = tk.Tk()
root.wm_title("Genetrator")
laby = Label(root, width=15, text="Year:")
enty = Entry(root, width=30)
but = Button(root, width=8, text="Okay", command=getList)
laby.grid(row=2,column=1,padx=5,pady=5)
enty.grid(row=2,column=2,columnspan=2,padx=10,pady=5)
but.grid(row=3,column=1,columnspan=3,pady=5)
root.mainloop()

def typ(a):
    for x, y in typs:
        if a == x:
            return y

def typno(a):
    for i in range(0, 12):
        if a == typs[i][0]:
            return i

def summer(ws,r,j):
    return float(ws.cell(row=r,column=j).value)+float(ws.cell(row=r+1,column=j).value)+float(ws.cell(row=r+2,column=j).value)+float(ws.cell(row=r+3,column=j).value)+float(ws.cell(row=r+4,column=j).value)+float(ws.cell(row=r+5,column=j).value)+float(ws.cell(row=r+6,column=j).value)+float(ws.cell(row=r+7,column=j).value)+float(ws.cell(row=r+8,column=j).value)+float(ws.cell(row=r+9,column=j).value)+float(ws.cell(row=r+10,column=j).value)

connection = pymysql.connect(host='localhost',
                             user='root',
                             password='',
                             db='datacollection',
                             cursorclass=pymysql.cursors.DictCursor)

wb = load_workbook('bin.xlsx')
ws1a = wb["T1a"]
ws2 = wb["T2"]
ws2a = wb["T2a"]
ws3 = wb["T3"]
ws3a = wb["T3a"]
ws4 = wb["T4"]
ws4a = wb["T4a"]
ws5 = wb["T5"]
ws5a = wb["T5a"]
ws6 = wb["T6"]
ws7 = wb["T7"]

r = 12
for s in range(0, 11):
    tye, x = typs[s]
    s += 1
    r += 1
    with connection.cursor() as cursor:
        sql = "SELECT * FROM `sheet1` where type = %s AND year=%s;"
        cursor.execute(sql, (tye,year))
        i = 0
        for row in cursor:
            ws2.cell(row=8+r+i,column=1).value=int(r+i-12)
            ws3.cell(row=7+r+i,column=1).value=int(r+i-12)
            ws4.cell(row=8+r+i,column=1).value=int(r+i-12)
            ws5.cell(row=8+r+i,column=1).value=int(r+i-12)
            ws6.cell(row=9+r+i,column=1).value=int(r+i-12)
            ws7.cell(row=9+r+i,column=1).value=int(r+i-12)
            ws2.cell(row=8+r+i,column=2).value=str(typ(tye))
            ws3.cell(row=7+r+i,column=2).value=str(typ(tye))
            ws4.cell(row=8+r+i,column=2).value=str(typ(tye))
            ws5.cell(row=8+r+i,column=2).value=str(typ(tye))
            ws6.cell(row=9+r+i,column=2).value=str(typ(tye))
            ws7.cell(row=9+r+i,column=2).value=str(typ(tye))
            ws2.cell(row=8+r+i,column=3).value=str(row['NOE'])
            ws3.cell(row=7+r+i,column=3).value=str(row['NOE'])
            ws4.cell(row=8+r+i,column=3).value=str(row['NOE'])
            ws5.cell(row=8+r+i,column=3).value=str(row['NOE'])
            ws6.cell(row=9+r+i,column=3).value=str(row['NOE'])
            ws7.cell(row=9+r+i,column=3).value=str(row['NOE'])
            i+=1

        sql = "SELECT * FROM `sheet2` where type = %s AND year=%s;"
        cursor.execute(sql,(tye,year))
        i=0
        for row in cursor:
            ws2.cell(row=8+r+i,column=4).value=row['1000IT']
            ws2.cell(row=8+r+i,column=5).value=row['1002IT']
            ws2.cell(row=8+r+i,column=6).value=float(row['1001IT'])+float(row['1007IT'])+float(row['1340IT'])
            ws2.cell(row=8+r+i,column=7).value=float(row['1110IT'])+float(row['1065IT'])
            ws2.cell(row=8+r+i,column=9).value=abs(float(row['1021IT'])) if float(row['1030IT'])>0 else 0
            ws2.cell(row=8+r+i,column=10).value=(float(row['1026IT'])+float(row['1028IT'])+float(row['1034IT'])+float(row['1036IT'])) if float(row['1030IT'])>0 else 0
            ws2.cell(row=8+r+i,column=11).value=row['1080IT']
            ws2.cell(row=8+r+i,column=12).value=row['1119IT']
            ws2.cell(row=8+r+i,column=15).value=row['1090IT']
            ws2.cell(row=8+r+i,column=17).value=row['1020IT']
            ws3.cell(row=7+r+i,column=14).value = abs(float(row['1030IT'])) if float(row['1030IT'])<0 else 0
            ws6.cell(row=9+r+i,column=4).value=row['1010IT']
            ws6.cell(row=9+r+i,column=5).value=float(row['1110IT'])+float(row['1065IT'])
            ws7.cell(row=9+r+i,column=4).value=row['1000IT']
            ws7.cell(row=9+r+i,column=10).value=row['1480IT']
            ws7.cell(row=9+r+i,column=11).value=row['1500IT']
            if (row['1500IT']>0):
                ws1a.cell(row=2*typno(tye)+14,column=5).value += 1
                ws1a.cell(row=2*typno(tye)+14,column=10).value += abs(float(row['1500IT']))
            if (row['1500IT']<0):
                ws1a.cell(row=2*typno(tye)+14,column=6).value += 1
                ws1a.cell(row=2*typno(tye)+14,column=11).value += abs(float(row['1500IT']))
            i+=1

        sql = "SELECT * FROM `sheet2a` where type = %s AND year=%s;"
        cursor.execute(sql,(tye,year))
        i=0
        for row in cursor:
            ws2.cell(row=8+r+i,column=8).value=float(row['1099IT'])+float(row['1060IT'])
            ws2.cell(row=8+r+i,column=13).value=row['1262IT']
            ws2.cell(row=8+r+i,column=14).value=row['1265IT']
            ws2.cell(row=8+r+i,column=16).value=float(row['1350IT'])+float(row['1360IT'])
            ws3.cell(row=7+r+i,column=4).value=row['1210IT']
            ws3.cell(row=7+r+i,column=5).value=float(row['1220IT'])+float(row['1221IT'])
            ws3.cell(row=7+r+i,column=6).value=row['1700IT']
            ws3.cell(row=7+r+i,column=7).value=row['1240IT']
            ws3.cell(row=7+r+i,column=8).value=row['1250IT']
            ws3.cell(row=7+r+i,column=9).value=row['1540IT']
            ws3.cell(row=7+r+i,column=10).value=row['1550IT']
            ws3.cell(row=7+r+i,column=11).value=row['1545IT']
            ws3.cell(row=7+r+i,column=12).value=row['1560IT']
            ws3.cell(row=7+r+i,column=13).value=row['1320IT']
            ws3.cell(row=7+r+i,column=15).value=float(row['1308IT'])+float(row['1620IT'])
            ws3.cell(row=7+r+i,column=16).value=row['1281IT']
            ws6.cell(row=9+r+i,column=7).value=float(row['1300IT'])-float(row['1265IT'])
            if (ws2.cell(row=8+r+i,column=1).value!=None):
                ws2.cell(row=8+r+i,column=18).value = float(ws2.cell(row=8+r+i,column=5).value)+float(ws2.cell(row=8+r+i,column=6).value)+float(ws2.cell(row=8+r+i,column=7).value)+float(ws2.cell(row=8+r+i,column=8).value)+float(ws2.cell(row=8+r+i,column=9).value)+float(ws2.cell(row=8+r+i,column=10).value)+float(ws2.cell(row=8+r+i,column=11).value)+float(ws2.cell(row=8+r+i,column=12).value)+float(ws2.cell(row=8+r+i,column=13).value)+float(ws2.cell(row=8+r+i,column=15).value)+float(ws2.cell(row=8+r+i,column=16).value)+float(ws2.cell(row=8+r+i,column=17).value)
                for j in range (4,19):
                    ws2.cell(row=typno(tye)+10,column=j).value+=float(ws2.cell(row=8+r+i,column=j).value)
            if (ws3.cell(row=7+r+i,column=1).value!=None):
                ws3.cell(row=7+r+i,column=17).value = float(ws3.cell(row=7+r+i,column=6).value)+float(ws3.cell(row=7+r+i,column=7).value)+float(ws3.cell(row=7+r+i,column=8).value)+float(ws3.cell(row=7+r+i,column=9).value)+float(ws3.cell(row=7+r+i,column=10).value)+float(ws3.cell(row=7+r+i,column=11).value)+float(ws3.cell(row=7+r+i,column=12).value)+float(ws3.cell(row=7+r+i,column=13).value)+float(ws3.cell(row=7+r+i,column=14).value)+float(ws3.cell(row=7+r+i,column=15).value)+float(ws3.cell(row=7+r+i,column=16).value)
                for j in range (4,18):
                    ws3.cell(row=typno(tye)+9,column=j).value+=float(ws3.cell(row=7+r+i,column=j).value)
            i+=1

        sql = "SELECT * FROM `sheet3` where type = %s AND year=%s;"
        cursor.execute(sql,(tye,year))
        i=0
        for row in cursor:
            ws4.cell(row=8+r+i,column=4).value=row['1410IT']
            ws4.cell(row=8+r+i,column=5).value=row['1411IT']
            ws4.cell(row=8+r+i,column=6).value=row['1403IT']
            ws4.cell(row=8+r+i,column=8).value=row['1402IT']
            ws5.cell(row=8+r+i,column=4).value=float(row['1407IT'])+float(row['1412IT'])
            ws5.cell(row=8+r+i,column=5).value=row['1409IT']
            ws5.cell(row=8+r+i,column=10).value=float(row['1413IT'])+float(row['1414IT'])+float(row['1415IT'])
            ws6.cell(row=9+r+i,column=6).value=row['1405IT']        
            ws7.cell(row=9+r+i,column=7).value=row['1405IT']
            ws7.cell(row=9+r+i,column=12).value=row['1409IT']
            i+=1 

        sql = "SELECT * FROM `sheet3a` where type = %s AND year=%s;"
        cursor.execute(sql,(tye,year))
        i=0
        for row in cursor:
            ws4.cell(row=8+r+i,column=7).value=abs(float(row['1435IT'])) if float(row['1435IT'])<0 else 0
            ws5.cell(row=8+r+i,column=6).value=(row['1435IT'] if float(row['1435IT'])>0 else 0)
            ws5.cell(row=8+r+i,column=7).value=row['1450IT']
            ws5.cell(row=8+r+i,column=8).value=row['1426IT']
            ws5.cell(row=8+r+i,column=9).value=float(row['1427IT'])+float(row['1465IT'])
            ws6.cell(row=9+r+i,column=8).value = float(row['1437IT'])/float(ws6.cell(row=9+r+i,column=7).value)*100 if (float(ws6.cell(row=9+r+i,column=7).value))!=0 else 0
            ws7.cell(row=9+r+i,column=9).value=row['1600IT']
            if (ws4.cell(row=8+r+i,column=1).value!=None):
                ws4.cell(row=8+r+i,column=9).value = float(ws4.cell(row=8+r+i,column=6).value)+float(ws4.cell(row=8+r+i,column=7).value)+float(ws4.cell(row=8+r+i,column=8).value)
                for j in range (4,10):
                        ws4.cell(row=typno(tye)+10,column=j).value+=float(ws4.cell(row=8+r+i,column=j).value)
            if (ws5.cell(row=8+r+i,column=1).value!=None):
                ws5.cell(row=8+r+i,column=11).value = float(ws5.cell(row=8+r+i,column=4).value)+float(ws5.cell(row=8+r+i,column=5).value)+float(ws5.cell(row=8+r+i,column=6).value)+float(ws5.cell(row=8+r+i,column=7).value)+float(ws5.cell(row=8+r+i,column=8).value)+float(ws5.cell(row=8+r+i,column=9).value)+float(ws5.cell(row=8+r+i,column=10).value)
                for j in range (4,12):
                        ws5.cell(row=typno(tye)+10,column=j).value+=float(ws5.cell(row=8+r+i,column=j).value)
            if (ws7.cell(row=9+r+i,column=1).value!=None):
                ws7.cell(row=9+r+i,column=8).value = ws5.cell(row=8+r+i,column=11).value
                ws7.cell(row=9+r+i,column=5).value = ws2.cell(row=8+r+i,column=18).value
                ws7.cell(row=9+r+i,column=6).value = ws3.cell(row=7+r+i,column=17).value
                for j in range (4,13):
                    ws7.cell(row=typno(tye)+11,column=j).value+=float(ws7.cell(row=9+r+i,column=j).value)
            i+=1

        sql = "SELECT * FROM `sheet5` where type = %s AND year=%s;"
        cursor.execute(sql,(tye,year))
        i=0
        for row in cursor:
            ws6.cell(row=9+r+i,column=9).value=row['4_8_1N']
            ws6.cell(row=9+r+i,column=10).value=row['4_8_2N']
            ws6.cell(row=9+r+i,column=11).value=row['4_8_3N']
            ws6.cell(row=9+r+i,column=12).value=float(row['4_8_4N'])+float(row['4_8_5N'])
            ws6.cell(row=9+r+i,column=13).value=row['4_8_8N']
            if (ws6.cell(row=9+r+i,column=1).value!=None):
                for j in range (4,8):
                    ws6.cell(row=typno(tye)+11,column=j).value+=float(ws6.cell(row=9+r+i,column=j).value)
                for j in range (9,14):
                    ws6.cell(row=typno(tye)+11,column=j).value+=float(ws6.cell(row=9+r+i,column=j).value)
                ws6.cell(row=typno(tye)+11,column=8).value=(float(ws7.cell(row=typno(tye)+11,column=9).value)+float(ws5.cell(row=typno(tye)+10,column=7).value)+float(ws5.cell(row=typno(tye)+10,column=9).value))/float(ws6.cell(row=typno(tye)+11,column=7).value)*100 if ws6.cell(row=typno(tye)+11,column=7).value!=0 else 0
            i+=1
        r=r+i-1

nume=summer(ws7,11,9)+summer(ws5,10,7)+summer(ws5,10,9)
denom=summer(ws6,11,7)
if denom>0:
    ws6.cell(row=10,column=8).value = nume/denom*100
else:
    ws6.cell(row=10,column=8).value = 0

for s in range(0, 11):
    tye, x = typs[s]
    for j in range (4,19):
        denom = summer(ws2,10,j)
        if denom>0:
            ws2a.cell(row=2*typno(tye)+11,column=j-1).value = float(ws2.cell(row=typno(tye)+10,column=j).value)
            ws2a.cell(row=2*typno(tye)+12,column=j-1).value = float(ws2.cell(row=typno(tye)+10,column=j).value)/denom*100
        else:
             ws2a.cell(row=2*typno(tye)+11,column=j-1).value = 0
             ws2a.cell(row=2*typno(tye)+12,column=j-1).value = 0
    for j in range (4,8):
        denom = summer(ws3,9,j)
        if denom>0:
            ws3a.cell(row=2*typno(tye)+9,column=j-1).value = float(ws3.cell(row=typno(tye)+9,column=j).value)
            ws3a.cell(row=2*typno(tye)+10,column=j-1).value = float(ws3.cell(row=typno(tye)+9,column=j).value)/denom*100
        else:
             ws3a.cell(row=2*typno(tye)+9,column=j-1).value = 0
             ws3a.cell(row=2*typno(tye)+10,column=j-1).value = 0            
    denom = summer(ws3,9,8)+summer(ws3,9,9)
    if denom>0:
        ws3a.cell(row=2*typno(tye)+9,column=7).value = float(ws3.cell(row=typno(tye)+9,column=8).value)+float(ws3.cell(row=typno(tye)+9,column=9).value)
        ws3a.cell(row=2*typno(tye)+10,column=7).value = (float(ws3.cell(row=typno(tye)+9,column=8).value)+float(ws3.cell(row=typno(tye)+9,column=9).value))/denom*100        
    else:
        ws3a.cell(row=2*typno(tye)+9,column=7).value = 0
        ws3a.cell(row=2*typno(tye)+10,column=7).value = 0
        
    for j in range (10,18):
        denom = summer(ws3,9,j)
        if denom>0:
            ws3a.cell(row=2*typno(tye)+9,column=j-2).value = float(ws3.cell(row=typno(tye)+9,column=j).value)
            ws3a.cell(row=2*typno(tye)+10,column=j-2).value = float(ws3.cell(row=typno(tye)+9,column=j).value)/denom*100
        else:
            ws3a.cell(row=2*typno(tye)+9,column=j-2).value = 0
            ws3a.cell(row=2*typno(tye)+10,column=j-2).value = 0            
    for j in range (4,10):
        denom = summer(ws4,10,j)
        if denom>0:
            ws4a.cell(row=2*typno(tye)+9,column=j-1).value = float(ws4.cell(row=typno(tye)+10,column=j).value)
            ws4a.cell(row=2*typno(tye)+10,column=j-1).value = float(ws4.cell(row=typno(tye)+10,column=j).value)/denom*100
        else:
            ws4a.cell(row=2*typno(tye)+9,column=j-1).value = 0
            ws4a.cell(row=2*typno(tye)+10,column=j-1).value = 0
    for j in range (4,12):
        denom = summer(ws5,10,j)
        if denom>0:
            ws5a.cell(row=2*typno(tye)+10,column=j-1).value = float(ws5.cell(row=typno(tye)+10,column=j).value)
            ws5a.cell(row=2*typno(tye)+11,column=j-1).value = float(ws5.cell(row=typno(tye)+10,column=j).value)/denom*100
        else:
            ws5a.cell(row=2*typno(tye)+10,column=j-1).value = 0
            ws5a.cell(row=2*typno(tye)+11,column=j-1).value = 0
    ws1a.cell(row=2*typno(tye)+14,column=7).value += float(ws2.cell(row=typno(tye)+10,column=4).value)+float(ws2.cell(row=typno(tye)+10,column=7).value)
    if (float(ws2.cell(row=typno(tye)+10,column=5).value)+float(ws2.cell(row=typno(tye)+10,column=6).value))>0:
        ws1a.cell(row=2*typno(tye)+14,column=9).value = float(ws2.cell(row=typno(tye)+10,column=7).value)/(float(ws2.cell(row=typno(tye)+10,column=5).value)+float(ws2.cell(row=typno(tye)+10,column=6).value))
    else:
        ws1a.cell(row=2*typno(tye)+14,column=9).value = 0
    ws1a.cell(row=2*typno(tye)+14,column=13).value = ws4.cell(row=typno(tye)+10,column=9).value

denom = summer(ws2,10,5)+summer(ws2,10,6)
if denom>0:
    ws1a.cell(row=37,column=9).value = summer(ws2,10,7)/denom
else:
    ws1a.cell(row=37,column=9).value = 0

if not os.path.exists("Output"):
    os.makedirs("Output")
wb.save("Output\\Report "+year+".xlsx")